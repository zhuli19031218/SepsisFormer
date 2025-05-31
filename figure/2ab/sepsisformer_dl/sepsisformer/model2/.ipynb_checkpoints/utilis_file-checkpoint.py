import os
import time
import pandas as pd

import openpyxl


def read_excel(name='./comparison_1.xlsx', row=1, column=1):
    wb = openpyxl.load_workbook(name)
    # print(type(wb))
    sheets = wb.get_sheet_names()
    # print(sheets)
    sheet = wb.get_sheet_by_name(sheets[0])
    return sheet.cell(row, column).value


def write_excel(name='./comparison_1.xlsx', row=1, column=1, v=' '):
    # print(name)
    wb = openpyxl.load_workbook(name)
    # print(type(wb))
    sheets = wb.get_sheet_names()
    # print(sheets)
    sheet = wb.get_sheet_by_name(sheets[0])

    sheet.cell(row, column).value = v
    wb.save(name)


def get_dir_name(epoch, lr):
    """
    :return: 生成一个以参数和时间戳命名的文件夹名,最终存放在log里
    """
    epoch = "epoch" + str(epoch)
    # lr = "lr" + str(lr)
    _time = str(time.strftime("%Y%m%d-%H-%M", time.localtime()))  # 获取当前epoch的运行时刻
    dir_name = r'_{}'.format(epoch)
    dir_name = _time + dir_name

    return dir_name


def mkdir(dir_name):
    """
    创建Logs文件夹，并以运行时间（年月日）+batch_size + epoch + Lr命名
    """
    if not os.path.exists('logs'):
        os.mkdir('logs')
    # if not os.path.exists('../model'):
    #     os.mkdir('../model')

    logs_name = os.path.join("logs", dir_name)
    if not os.path.exists(logs_name):
        os.mkdir(logs_name)

    # model_name = os.path.join("../model", dir_name)
    # if not os.path.exists(model_name):
    #     os.mkdir(model_name)

    return logs_name  # , model_name


def save_file(data: pd.DataFrame, dir_name: str, file_name):
    """
    将pd格式数据写入指定文件夹
    """
    name = os.path.join(dir_name, file_name)  # name表示文件夹路径
    data.to_csv(name, index=False, header=False)  # 将pd格式数据写入’.CSV‘表格文件
    print(file_name, "save success!")
